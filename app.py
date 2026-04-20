import os
import json
import base64
import pickle
from datetime import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import numpy as np
import cv2
import face_recognition
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__, static_folder='static')
CORS(app)

ENCODINGS_FILE = 'face_encodings.pkl'
ATTENDANCE_FILE = 'attendance.xlsx'
ATTENDANCE_LOG = []  # in-memory log for current session

def load_encodings():
    if os.path.exists(ENCODINGS_FILE):
        with open(ENCODINGS_FILE, 'rb') as f:
            return pickle.load(f)
    return {}

def save_encodings(data):
    with open(ENCODINGS_FILE, 'wb') as f:
        pickle.dump(data, f)

def decode_image(b64_string):
    if ',' in b64_string:
        b64_string = b64_string.split(',')[1]
    img_bytes = base64.b64decode(b64_string)
    np_arr = np.frombuffer(img_bytes, np.uint8)
    img = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
    return cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

def build_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(bold=True, color="E8D5A3", size=12, name="Arial")
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Name", "Date", "Time", "Status"]
    col_widths = [5, 25, 15, 15, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30

    alt_fill = PatternFill("solid", fgColor="F5F0E8")
    norm_fill = PatternFill("solid", fgColor="FFFFFF")

    seen_today = {}
    for i, record in enumerate(ATTENDANCE_LOG, 1):
        fill = alt_fill if i % 2 == 0 else norm_fill
        row_data = [i, record['name'], record['date'], record['time'], record['status']]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            cell.font = Font(name="Arial", size=11)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = "Attendance Summary"
    ws2['A1'].font = Font(bold=True, size=14, name="Arial", color="1A1A2E")
    ws2['A3'] = "Total Records"
    ws2['B3'] = len(ATTENDANCE_LOG)
    ws2['A4'] = "Generated"
    ws2['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    unique = {}
    for r in ATTENDANCE_LOG:
        unique[r['name']] = unique.get(r['name'], 0) + 1
    row = 6
    ws2['A5'] = "Name"
    ws2['B5'] = "Check-ins"
    ws2['A5'].font = Font(bold=True, name="Arial")
    ws2['B5'].font = Font(bold=True, name="Arial")
    for name, count in unique.items():
        ws2[f'A{row}'] = name
        ws2[f'B{row}'] = count
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/register', methods=['POST'])
def register():
    data = request.json
    name = data.get('name', '').strip()
    image_b64 = data.get('image')

    if not name or not image_b64:
        return jsonify({'success': False, 'error': 'Name and image required'})

    try:
        img = decode_image(image_b64)
        encodings = face_recognition.face_encodings(img)

        if not encodings:
            return jsonify({'success': False, 'error': 'No face detected. Please try again with better lighting.'})

        db = load_encodings()
        db[name] = encodings[0]
        save_encodings(db)

        return jsonify({'success': True, 'message': f'{name} registered successfully!', 'total': len(db)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/recognize', methods=['POST'])
def recognize():
    data = request.json
    image_b64 = data.get('image')

    if not image_b64:
        return jsonify({'success': False, 'error': 'No image provided'})

    try:
        img = decode_image(image_b64)
        db = load_encodings()

        if not db:
            return jsonify({'success': False, 'error': 'No registered faces. Please register first.'})

        face_locs = face_recognition.face_locations(img)
        face_encs = face_recognition.face_encodings(img, face_locs)

        if not face_encs:
            return jsonify({'recognized': [], 'unknown': 0})

        known_names = list(db.keys())
        known_encs = list(db.values())
        results = []

        for enc, loc in zip(face_encs, face_locs):
            distances = face_recognition.face_distance(known_encs, enc)
            best_idx = np.argmin(distances)
            if distances[best_idx] < 0.5:
                name = known_names[best_idx]
                now = datetime.now()
                record = {
                    'name': name,
                    'date': now.strftime('%Y-%m-%d'),
                    'time': now.strftime('%H:%M:%S'),
                    'status': 'Present',
                    'confidence': round((1 - distances[best_idx]) * 100, 1)
                }
                # Avoid duplicate entries within same minute
                already = any(
                    r['name'] == name and r['date'] == record['date'] and r['time'][:5] == record['time'][:5]
                    for r in ATTENDANCE_LOG
                )
                if not already:
                    ATTENDANCE_LOG.append(record)
                results.append({'name': name, 'location': loc, 'confidence': record['confidence']})
            else:
                results.append({'name': 'Unknown', 'location': loc, 'confidence': 0})

        unknown_count = sum(1 for r in results if r['name'] == 'Unknown')
        recognized = [r for r in results if r['name'] != 'Unknown']
        return jsonify({'recognized': recognized, 'unknown': unknown_count})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/attendance', methods=['GET'])
def get_attendance():
    return jsonify({'records': ATTENDANCE_LOG, 'total': len(ATTENDANCE_LOG)})

@app.route('/registered', methods=['GET'])
def get_registered():
    db = load_encodings()
    return jsonify({'names': list(db.keys()), 'total': len(db)})

@app.route('/delete/<name>', methods=['DELETE'])
def delete_person(name):
    db = load_encodings()
    if name in db:
        del db[name]
        save_encodings(db)
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Person not found'})

@app.route('/clear_attendance', methods=['POST'])
def clear_attendance():
    global ATTENDANCE_LOG
    ATTENDANCE_LOG = []
    return jsonify({'success': True})

@app.route('/export', methods=['GET'])
def export():
    excel_data = build_excel()
    filename = f"attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    os.makedirs('static', exist_ok=True)
    app.run(debug=True, port=5000)
